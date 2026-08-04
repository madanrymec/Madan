import io
import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from flask import Flask
from flask_apscheduler import APScheduler
import mysql.connector

# ----------------------------------------------------------------------
# DATABASE CONNECTION
# ----------------------------------------------------------------------
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="password",
        database="your_database_name"
    )

# ----------------------------------------------------------------------
# HELPER PARSING UTILITIES
# ----------------------------------------------------------------------
def parse_trailing_minus_decimal(val):
    """Handles SAP trailing minus signs (e.g., '123.45-' -> '-123.45')."""
    if val is None or val == "":
        return Decimal('0.00')
    val_str = str(val).strip().replace(',', '')
    if val_str.endswith('-'):
        val_str = '-' + val_str[:-1]
    try:
        return Decimal(val_str)
    except (InvalidOperation, ValueError):
        return Decimal('0.00')


def parse_safe_date(val):
    """Converts Excel date objects, datetimes, or date strings safely into YYYY-MM-DD."""
    if val is None or val == "":
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    
    val_str = str(val).strip()
    if not val_str:
        return None
        
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y%m%d'):
        try:
            return datetime.strptime(val_str[:10], fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


# ----------------------------------------------------------------------
# FLASK & SCHEDULER ENGINE SETUP
# ----------------------------------------------------------------------
app = Flask(__name__)

class SchedulerConfig:
    SCHEDULER_API_ENABLED = False
    SCHEDULER_EXECUTORS = {"default": {"type": "threadpool", "max_workers": 1}}

app.config.from_object(SchedulerConfig())
scheduler = APScheduler()
scheduler.init_app(app)
scheduler.start()

print("[SCHEDULER ENGINE] Background task runner initialized successfully.")


# ----------------------------------------------------------------------
# JOB 1: COLLECTIONS SET SYNC (HOURLY)
# ----------------------------------------------------------------------
@scheduler.task('interval', id='collections_set_sync_job', hours=1, misfire_grace_time=900)
def automated_collections_set_job():
    print("\n" + "═" * 60)
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] AUTOMATED COLLECTIONS SET SYNC TRIGGERED")
    print("═" * 60)

    with app.app_context():
        sap_url = "https://kecapi.kirloskar-electric.com/sap/opu/odata/sap/ZCM_ARR_SRV/ZCM_ARRSet/?$format=xlsx"
        username = "corpodata"
        password = "Kec12345"

        try:
            # --- STEP 1: NETWORK REQUEST TO SAP ODATA ---
            print("[COLLECTIONS-SET STEP 1/4] Establishing secure connection link to live SAP systems...")
            response = requests.get(sap_url, auth=HTTPBasicAuth(username, password), timeout=360)
            print(f"--> Server response received: {response.status_code}")

            if response.status_code != 200:
                print("[CRITICAL AUTOMATED JOB ERROR] Live SAP stream network access drop. Job interrupted.")
                return

            # --- STEP 2: WORKBOOK STREAM PARSING ---
            print("[COLLECTIONS-SET STEP 2/4] Parsing streaming binary sheet data directly out of RAM...")
            excel_file_bytes = io.BytesIO(response.content)
            wb = openpyxl.load_workbook(excel_file_bytes, data_only=True)
            ws = wb.active

            total_rows_found = ws.max_row - 1
            print(f"--> Stream processed. {total_rows_found} live rows found inside the current stream pool.")

            if total_rows_found <= 0:
                print("[INFO] Incoming data grid array is empty. Closing interval.")
                return

            # --- STEP 3: DETECT ACTIVE DATE SCOPE AND PURGE TARGET BLOCK ---
            print("[COLLECTIONS-SET STEP 3/4] Initializing database transaction context...")
            conn = get_db_connection()
            cursor = conn.cursor()

            # Inspect the first data row (row 2) to identify AS_ON_Date
            first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

            if first_row and first_row[0] is not None:
                target_date = parse_safe_date(first_row[0])

                if target_date:
                    print(f"--> Target scope detected: AS_ON_Date {target_date}")
                    print(f"--> Clearing existing database entries for current target scope ({target_date})...")

                    delete_query = "DELETE FROM collections_set WHERE AS_ON_Date = %s"
                    cursor.execute(delete_query, (target_date,))
                    print(f"--> Target scope purge complete. Removed {cursor.rowcount} temporary rows.")
                else:
                    print("[WARNING] Could not parse date bounds safely from row data. Proceeding without safety purge.")
            else:
                print("[WARNING] Spreadsheet data row structure invalid. Proceeding without safety purge.")

            # --- STEP 4: RECORD PROCESSING ENGINE LOOP ---
            print("[COLLECTIONS-SET STEP 4/4] Executing fresh insertion batch loop...")
            insert_query = """
                INSERT INTO collections_set (
                    AS_ON_Date, Company_Code, Customer_number, Assignment, Fiscal_Year, Document_Number,
                    Line_item, Billing_Document, Posting_Key, Document_Type, Profit_Center, Name1, City,
                    Invoice_Date, Amount_in_LC, Balance, Balance_In_Lakhs, Baseline_Payment_Dte, Due_Outstanding,
                    Invoice_Age, Day_Difference, Aging_Bucket, Invoice_Text, Collected_Amount, Pay_Date,
                    Cheque_Number, Balance_Text, Realization_Days, Branch, Description, Risk_Category,
                    Special_G_L_ind, Special_Transaction, Sales_Document, Sales_Document_Date, Purchase_order_no,
                    Purchase_order_date, Terms_of_Payment, Terms_of_Payment_Key, Region, Product_group, Material,
                    LR_Number, LR_Date, Customer_classification, Description3, Unitwise, Created_On, Created_By,
                    Changed_On, Change_At, Change_By, Sales_Engineer_ID, Sales_Engineer_Name, Sales_Engineer_Email_ID,
                    Branch_Manager_ID, Branch_Manager_Name, Branch_Manager_Email_ID, CFD_ID, CFD_Name,
                    CFD_Email_ID, Ignore_Invoice, No_of_Iterations, Cancelled_Bill_Doc, Posting_Date, Advance_Amount,
                    CURRENT_BALANCE, System_Date, Sorting_Key
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """

            inserted_counter = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue

                # Parse Trailing Minus for Balance (Index 15)
                balance_val = parse_trailing_minus_decimal(row[15]) if len(row) > 15 else Decimal('0.00')

                cursor.execute(insert_query, (
                    parse_safe_date(row[0]) if len(row) > 0 else None,
                    str(row[1]).strip() if len(row) > 1 and row[1] is not None else None,
                    str(row[2]).strip() if len(row) > 2 and row[2] is not None else None,
                    str(row[3]).strip() if len(row) > 3 and row[3] is not None else None,
                    str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
                    str(row[5]).strip() if len(row) > 5 and row[5] is not None else None,
                    str(row[6]).strip() if len(row) > 6 and row[6] is not None else None,
                    str(row[7]).strip() if len(row) > 7 and row[7] is not None else None,
                    str(row[8]).strip() if len(row) > 8 and row[8] is not None else None,
                    str(row[9]).strip() if len(row) > 9 and row[9] is not None else None,
                    str(row[10]).strip() if len(row) > 10 and row[10] is not None else None,
                    str(row[11]).strip() if len(row) > 11 and row[11] is not None else None,
                    str(row[12]).strip() if len(row) > 12 and row[12] is not None else None,
                    parse_safe_date(row[13]) if len(row) > 13 else None,
                    parse_trailing_minus_decimal(row[14]) if len(row) > 14 else Decimal('0.00'),
                    balance_val,
                    parse_trailing_minus_decimal(row[16]) if len(row) > 16 else Decimal('0.00'),
                    parse_safe_date(row[17]) if len(row) > 17 else None,
                    str(row[18]).strip() if len(row) > 18 and row[18] is not None else None,
                    str(row[19]).strip() if len(row) > 19 and row[19] is not None else None,
                    str(row[20]).strip() if len(row) > 20 and row[20] is not None else None,
                    str(row[21]).strip() if len(row) > 21 and row[21] is not None else None,
                    str(row[22]).strip() if len(row) > 22 and row[22] is not None else None,
                    parse_trailing_minus_decimal(row[23]) if len(row) > 23 else Decimal('0.00'),
                    parse_safe_date(row[24]) if len(row) > 24 else None,
                    str(row[25]).strip() if len(row) > 25 and row[25] is not None else None,
                    str(row[26]).strip() if len(row) > 26 and row[26] is not None else None,
                    str(row[27]).strip() if len(row) > 27 and row[27] is not None else None,
                    str(row[28]).strip() if len(row) > 28 and row[28] is not None else None,
                    str(row[29]).strip() if len(row) > 29 and row[29] is not None else None,
                    str(row[30]).strip() if len(row) > 30 and row[30] is not None else None,
                    str(row[31]).strip() if len(row) > 31 and row[31] is not None else None,
                    str(row[32]).strip() if len(row) > 32 and row[32] is not None else None,
                    str(row[33]).strip() if len(row) > 33 and row[33] is not None else None,
                    parse_safe_date(row[34]) if len(row) > 34 else None,
                    str(row[35]).strip() if len(row) > 35 and row[35] is not None else None,
                    parse_safe_date(row[36]) if len(row) > 36 else None,
                    str(row[37]).strip() if len(row) > 37 and row[37] is not None else None,
                    str(row[38]).strip() if len(row) > 38 and row[38] is not None else None,
                    str(row[39]).strip() if len(row) > 39 and row[39] is not None else None,
                    str(row[40]).strip() if len(row) > 40 and row[40] is not None else None,
                    str(row[41]).strip() if len(row) > 41 and row[41] is not None else None,
                    str(row[42]).strip() if len(row) > 42 and row[42] is not None else None,
                    parse_safe_date(row[43]) if len(row) > 43 else None,
                    str(row[44]).strip() if len(row) > 44 and row[44] is not None else None,
                    str(row[45]).strip() if len(row) > 45 and row[45] is not None else None,
                    str(row[46]).strip() if len(row) > 46 and row[46] is not None else None,
                    parse_safe_date(row[47]) if len(row) > 47 else None,
                    str(row[48]).strip() if len(row) > 48 and row[48] is not None else None,
                    parse_safe_date(row[49]) if len(row) > 49 else None,
                    str(row[50]).strip() if len(row) > 50 and row[50] is not None else None,
                    str(row[51]).strip() if len(row) > 51 and row[51] is not None else None,
                    str(row[52]).strip() if len(row) > 52 and row[52] is not None else None,
                    str(row[53]).strip() if len(row) > 53 and row[53] is not None else None,
                    str(row[54]).strip() if len(row) > 54 and row[54] is not None else None,
                    str(row[55]).strip() if len(row) > 55 and row[55] is not None else None,
                    str(row[56]).strip() if len(row) > 56 and row[56] is not None else None,
                    str(row[57]).strip() if len(row) > 57 and row[57] is not None else None,
                    str(row[58]).strip() if len(row) > 58 and row[58] is not None else None,
                    str(row[59]).strip() if len(row) > 59 and row[59] is not None else None,
                    str(row[60]).strip() if len(row) > 60 and row[60] is not None else None,
                    str(row[61]).strip() if len(row) > 61 and row[61] is not None else None,
                    str(row[62]).strip() if len(row) > 62 and row[62] is not None else None,
                    str(row[63]).strip() if len(row) > 63 and row[63] is not None else None,
                    parse_safe_date(row[64]) if len(row) > 64 else None,
                    parse_trailing_minus_decimal(row[65]) if len(row) > 65 else Decimal('0.00'),
                    parse_trailing_minus_decimal(row[66]) if len(row) > 66 else Decimal('0.00'),
                    parse_safe_date(row[67]) if len(row) > 67 else None,
                    str(row[68]).strip() if len(row) > 68 and row[68] is not None else None
                ))
                inserted_counter += 1

            conn.commit()

            print("═" * 60)
            print(f"BACKGROUND SCHEDULE JOB PIPELINE METRICS RECONCILED:")
            print(f"--> Total Fresh Rows Overwritten and Saved: {inserted_counter}")
            print("═" * 60 + "\n")

        except Exception as e:
            print(f"\n[BACKGROUND SCHEDULER SYSTEM EXCEPTION CRASH]: {e}")
            if 'conn' in locals(): conn.rollback()
        finally:
            if 'cursor' in locals(): cursor.close()
            if 'conn' in locals(): conn.close()


# ----------------------------------------------------------------------
# JOB 2: COLLECTIONS DATA SYNC (HOURLY)
# ----------------------------------------------------------------------
@scheduler.task('interval', id='collections_data_sync_job', hours=1, misfire_grace_time=900)
def automated_collections_data_job():
    print("\n" + "═" * 60)
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] AUTOMATED COLLECTIONS DATA SYNC TRIGGERED")
    print("═" * 60)

    with app.app_context():
        sap_url = "https://kecapi.kirloskar-electric.com/sap/opu/odata/sap/ZCOLL_DOWNLOAD_SRV/ZCOLLSet?$format=xlsx"
        username = "corpodata"
        password = "Kec12345"

        try:
            # --- STEP 1: NETWORK REQUEST TO SAP ODATA ---
            print("[COLLECTIONS-DATA STEP 1/4] Establishing secure connection link to live SAP systems...")
            response = requests.get(sap_url, auth=HTTPBasicAuth(username, password), timeout=360)
            print(f"--> Server response received: {response.status_code}")

            if response.status_code != 200:
                print("[CRITICAL AUTOMATED JOB ERROR] Live SAP stream network access drop. Job interrupted.")
                return

            # --- STEP 2: WORKBOOK STREAM PARSING ---
            print("[COLLECTIONS-DATA STEP 2/4] Parsing streaming binary sheet data directly out of RAM...")
            excel_file_bytes = io.BytesIO(response.content)
            wb = openpyxl.load_workbook(excel_file_bytes, data_only=True)
            ws = wb.active

            total_rows_found = ws.max_row - 1
            print(f"--> Stream processed. {total_rows_found} live rows found inside the current stream pool.")

            if total_rows_found <= 0:
                print("[INFO] Incoming data grid array is empty. Closing interval.")
                return

            # --- STEP 3: DETECT ACTIVE DATE SCOPE AND PURGE TARGET BLOCK ---
            print("[COLLECTIONS-DATA STEP 3/4] Initializing database transaction context...")
            conn = get_db_connection()
            cursor = conn.cursor()

            # Inspect the first data row (row 2) to identify AS_ON_Date
            first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

            if first_row and first_row[0] is not None:
                target_date = parse_safe_date(first_row[0])

                if target_date:
                    print(f"--> Target scope detected: AS_ON_Date {target_date}")
                    print(f"--> Clearing existing database entries for current target scope ({target_date})...")

                    delete_query = "DELETE FROM collections_data WHERE AS_ON_Date = %s"
                    cursor.execute(delete_query, (target_date,))
                    print(f"--> Target scope purge complete. Removed {cursor.rowcount} temporary rows.")
                else:
                    print("[WARNING] Could not parse date bounds safely from row data. Proceeding without safety purge.")
            else:
                print("[WARNING] Spreadsheet data row structure invalid. Proceeding without safety purge.")

            # --- STEP 4: RECORD PROCESSING ENGINE LOOP ---
            print("[COLLECTIONS-DATA STEP 4/4] Executing fresh insertion batch loop...")
            insert_query = """
                INSERT INTO collections_data (
                    AS_ON_Date, Document_Number, Document_Date, Posting_Date, Cust_Code, Cust_Name,
                    Doc_Type, Cheque_No, Cheque_Dt, Cheque_Amt, Bank_Code, Bank_Desc, Unit_Code, Posting_Key,
                    Assignment_No, Invoice_Number, Invoice_Date, Realised_Amount, Cleared_Amount, Profit_Centre,
                    Prod_GRP_Code, Narration, Coll_BR_Code, Coll_BR_DESC, EXP_GLCD, EXP_PK, EXP_AMT, IC_ALT_DOC_NO,
                    ALT_CO_CODE, PAYMENT_TYPE, SAP_COLL_AMT, PROFIT_CENTRE_1, UPDT_FLAG
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """

            inserted_counter = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue

                # Parse Trailing Minus for SAP_COLL_AMT (Index 30)
                sap_coll_amt = parse_trailing_minus_decimal(row[30]) if len(row) > 30 else Decimal('0.00')

                cursor.execute(insert_query, (
                    parse_safe_date(row[0]) if len(row) > 0 else None,
                    str(row[1]).strip() if len(row) > 1 and row[1] is not None else None,
                    parse_safe_date(row[2]) if len(row) > 2 else None,
                    parse_safe_date(row[3]) if len(row) > 3 else None,
                    str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
                    str(row[5]).strip() if len(row) > 5 and row[5] is not None else None,
                    str(row[6]).strip() if len(row) > 6 and row[6] is not None else None,
                    str(row[7]).strip() if len(row) > 7 and row[7] is not None else None,
                    parse_safe_date(row[8]) if len(row) > 8 else None,
                    parse_trailing_minus_decimal(row[9]) if len(row) > 9 else Decimal('0.00'),
                    str(row[10]).strip() if len(row) > 10 and row[10] is not None else None,
                    str(row[11]).strip() if len(row) > 11 and row[11] is not None else None,
                    str(row[12]).strip() if len(row) > 12 and row[12] is not None else None,
                    str(row[13]).strip() if len(row) > 13 and row[13] is not None else None,
                    str(row[14]).strip() if len(row) > 14 and row[14] is not None else None,
                    str(row[15]).strip() if len(row) > 15 and row[15] is not None else None,
                    parse_safe_date(row[16]) if len(row) > 16 else None,
                    parse_trailing_minus_decimal(row[17]) if len(row) > 17 else Decimal('0.00'),
                    parse_trailing_minus_decimal(row[18]) if len(row) > 18 else Decimal('0.00'),
                    str(row[19]).strip() if len(row) > 19 and row[19] is not None else None,
                    str(row[20]).strip() if len(row) > 20 and row[20] is not None else None,
                    str(row[21]).strip() if len(row) > 21 and row[21] is not None else None,
                    str(row[22]).strip() if len(row) > 22 and row[22] is not None else None,
                    str(row[23]).strip() if len(row) > 23 and row[23] is not None else None,
                    str(row[24]).strip() if len(row) > 24 and row[24] is not None else None,
                    str(row[25]).strip() if len(row) > 25 and row[25] is not None else None,
                    parse_trailing_minus_decimal(row[26]) if len(row) > 26 else Decimal('0.00'),
                    str(row[27]).strip() if len(row) > 27 and row[27] is not None else None,
                    str(row[28]).strip() if len(row) > 28 and row[28] is not None else None,
                    str(row[29]).strip() if len(row) > 29 and row[29] is not None else None,
                    sap_coll_amt,
                    str(row[31]).strip() if len(row) > 31 and row[31] is not None else None,
                    str(row[32]).strip() if len(row) > 32 and row[32] is not None else None
                ))
                inserted_counter += 1

            conn.commit()

            print("═" * 60)
            print(f"BACKGROUND SCHEDULE JOB PIPELINE METRICS RECONCILED:")
            print(f"--> Total Fresh Rows Overwritten and Saved: {inserted_counter}")
            print("═" * 60 + "\n")

        except Exception as e:
            print(f"\n[BACKGROUND SCHEDULER SYSTEM EXCEPTION CRASH]: {e}")
            if 'conn' in locals(): conn.rollback()
        finally:
            if 'cursor' in locals(): cursor.close()
            if 'conn' in locals(): conn.close()


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)